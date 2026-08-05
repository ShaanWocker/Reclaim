"""
Builds a Power BI-ready Excel workbook from the message data.

Each sheet is written as a proper named Excel Table (not just a raw
range of cells) -- that's what lets Power BI's "Get Data > Excel
Workbook" pick up clean, typed columns and a ready-made query
automatically, instead of guessing at where a plain range starts and
stops.

Three tables:
  - MessagesTable: one row per email (sender, date, size, ID)
  - SendersTable:  totals per sender
  - MonthlyTable:  totals per calendar month, for time-series charts
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = Font(name="Calibri", size=11)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)


def _write_table(ws, df, table_name, date_formats=None, start_row=1):
    """Write a DataFrame into a worksheet as a named Excel Table."""
    date_formats = date_formats or {}
    columns = list(df.columns)
    n_rows = len(df)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, (col_name, value) in enumerate(zip(columns, row), start=1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT
            if col_name in date_formats and value is not None:
                cell.number_format = date_formats[col_name]

    if n_rows == 0:
        # Excel Tables need at least one data row -- pad with a blank one.
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=start_row + 1, column=col_idx, value=None)
        n_rows = 1

    last_col = get_column_letter(len(columns))
    last_row = start_row + n_rows
    table = Table(displayName=table_name, ref=f"A{start_row}:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for col_idx, col_name in enumerate(columns, start=1):
        longest = df.iloc[:, col_idx - 1].map(lambda v: len(str(v))).max() if len(df) else 0
        width = max(len(str(col_name)), longest)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 50)


def build_powerbi_workbook(df):
    """
    df: the (already filtered) messages DataFrame from the dashboard,
    with columns id, from, date, size_bytes, size_mb, date_parsed.

    Returns the .xlsx file as bytes, ready for st.download_button.
    """
    messages = df[["from", "date_parsed", "size_mb", "size_bytes", "id"]].copy()
    messages.columns = ["Sender", "Date", "SizeMB", "SizeBytes", "MessageID"]
    messages["Date"] = messages["Date"].dt.tz_localize(None)
    messages["SizeMB"] = messages["SizeMB"].round(3)

    senders = (
        df.groupby("from")
        .agg(TotalMB=("size_mb", "sum"), MessageCount=("id", "count"))
        .reset_index()
        .rename(columns={"from": "Sender"})
        .sort_values("TotalMB", ascending=False)
    )
    senders["TotalMB"] = senders["TotalMB"].round(3)
    senders["AvgMB"] = (senders["TotalMB"] / senders["MessageCount"]).round(3)

    monthly = df.dropna(subset=["date_parsed"]).copy()
    monthly["Month"] = (
        monthly["date_parsed"].dt.tz_localize(None).dt.to_period("M").dt.to_timestamp()
    )
    monthly = (
        monthly.groupby("Month")
        .agg(TotalMB=("size_mb", "sum"), MessageCount=("id", "count"))
        .reset_index()
        .sort_values("Month")
    )
    monthly["TotalMB"] = monthly["TotalMB"].round(3)

    wb = Workbook()

    ws_messages = wb.active
    ws_messages.title = "Messages"
    _write_table(ws_messages, messages, "MessagesTable", date_formats={"Date": "yyyy-mm-dd hh:mm"})

    ws_senders = wb.create_sheet("Senders")
    _write_table(ws_senders, senders, "SendersTable")

    ws_monthly = wb.create_sheet("Monthly")
    _write_table(ws_monthly, monthly, "MonthlyTable", date_formats={"Month": "yyyy-mm"})

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
